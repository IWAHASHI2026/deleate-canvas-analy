import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ライバル商品比較表"

# === スタイル定義 ===
title_font = Font(name="Meiryo UI", size=14, bold=True, color="FFFFFF")
title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

header_font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
rival_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
our_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
category_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
category_font = Font(name="Meiryo UI", size=10, bold=True)

cell_font = Font(name="Meiryo UI", size=10)
highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 当社優位
weak_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # 当社劣位

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# === 列幅設定 ===
col_widths = [22, 30, 30, 30, 30, 35]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# === タイトル行 ===
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "キャンバスプリント ライバル商品 vs 当社商品 マーケティング比較表"
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = center_align
ws.row_dimensions[1].height = 40

# === ヘッダー行 ===
headers = [
    ("比較項目", category_fill, Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")),
    ("ライバル①\n正方形キャンバス\n(izumidaishi)", rival_fill, header_font),
    ("ライバル②\n長方形キャンバス\n(izumidaishi)", rival_fill, header_font),
    ("当社①\nキャンバスプリント\n全面印刷 (dtu301)", our_fill, header_font),
    ("当社②\nキャンバスプリント\n(dtu300)", our_fill, header_font),
    ("マーケティング分析\n（当社の優位性/課題）", PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"), Font(name="Meiryo UI", size=10, bold=True)),
]

ws.row_dimensions[2].height = 55
for col, (text, fill, font) in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=text)
    cell.font = font
    cell.fill = fill
    cell.alignment = center_align
    cell.border = thin_border

# === 比較データ ===
# (カテゴリ, 項目, ライバル①, ライバル②, 当社①, 当社②, 分析, 当社有利かどうか)
# highlight: "good"=当社優位, "bad"=当社劣位, ""=同等
data = [
    ("価格戦略", "最低価格", "1,000円", "1,000円", "1,980円", "1,980円",
     "【課題】ライバルは1,000円ポッキリ訴求。当社は約2倍の価格差。価格競争力で劣後。", "bad"),

    ("", "価格帯上限", "4,800円", "4,980円", "3,080円", "3,080円",
     "当社は上限が低く、高額商品の展開が少ない。客単価向上の余地あり。", ""),

    ("", "価格訴求力", "「1000円ポッキリ」\nインパクト大", "「1000円ポッキリ」\nインパクト大", "特になし", "特になし",
     "【課題】ライバルの「1000円ポッキリ」は極めて強いキャッチコピー。当社にも価格訴求の工夫が必要。", "bad"),

    ("商品力", "サイズ展開", "正方形4種\n150~300mm", "長方形4種\nS~LL (150x200~300x400mm)", "3種\nS/M/L (10x13.3~20x26.6cm)", "3種\nS/M/L (10x13.3~20x26.6cm)",
     "【課題】ライバルは正方形・長方形合計8種。当社は3種のみ。サイズバリエーション拡充が必要。", "bad"),

    ("", "付属品", "なし", "なし", "ミニアクリルスタンド付", "ミニアクリルスタンド付",
     "【優位】アクリルスタンド付属は差別化要素。壁掛け不要で置ける利便性。", "good"),

    ("", "素材・品質", "布地・UVプリント\n厚さ15mm", "木製・紙\nUV印刷・厚さ15mm", "布製キャンバス\nUV加工", "キャンバス\nUV特殊プリント加工",
     "素材はほぼ同等。UV加工は共通。大きな差別化要因にはなりにくい。", ""),

    ("", "セット販売", "2枚セットあり\n(300SQ×2 ¥3,000)", "2~6枚セット\n複数枚割引あり", "なし", "なし",
     "【課題】ライバルはセット販売で客単価UP。当社もセット販売導入を検討すべき。", "bad"),

    ("納期・配送", "納期", "5営業日以内", "5営業日以内", "当日発送\n(12時まで注文)", "当日発送\n(12時まで注文)",
     "【優位】当社の当日発送は圧倒的な強み。急ぎのギフト需要を獲得可能。", "good"),

    ("", "送料", "無料", "無料", "無料（送料込み）", "無料（送料込み）",
     "送料は同等。差別化要因にはならない。", ""),

    ("", "保証", "記載なし", "記載なし", "90日間保証・返品対応", "記載なし",
     "【優位】90日保証は安心感を訴求できる。全商品に展開すべき。", "good"),

    ("カスタマイズ", "デザインツール", "DesignSoft\n（無料提供）", "デザイン確認・修正対応", "スマホアプリで\nレイアウト・プレビュー", "スマホアプリで\nレイアウト設計",
     "ほぼ同等。当社はスマホ対応が強み。「スマホで簡単」を訴求ポイントに。", "good"),

    ("", "文字入れ対応", "対応", "自由に文字設定可能", "文字入れ・名入れ対応", "テキスト・名入れ対応",
     "同等。差別化要因にはならない。", ""),

    ("", "用途提案", "特に記載なし", "特に記載なし", "結婚祝い・出産祝い\n記念日・ペット写真", "結婚祝い・出産祝い\n記念日・ペット写真",
     "【優位】当社はギフト用途の提案が充実。ギフト需要の取り込みに有利。", "good"),

    ("信頼性", "レビュー評価", "4.93/5.0", "4.90/5.0", "未確認", "未確認",
     "【課題】ライバルは高評価レビュー多数。当社もレビュー獲得施策が必要。", "bad"),

    ("", "レビュー件数", "710件", "495件", "未確認", "未確認",
     "【課題】ライバルは合計1,200件超のレビュー。社会的証明で大きく劣後。", "bad"),

    ("ブランド", "ブランド訴求", "ノーブランド", "ノーブランド", "名入れ屋\n（専門店ブランド）", "名入れ屋\n（専門店ブランド）",
     "【優位】「名入れ屋」は専門店としてのブランド力あり。信頼感の訴求に活用可能。", "good"),
]

row = 3
current_category = ""
for item in data:
    category, label, r1, r2, o1, o2, analysis, advantage = item

    # カテゴリ行
    if category and category != current_category:
        current_category = category
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cat_cell = ws.cell(row=row, column=1, value=f"■ {category}")
        cat_cell.font = category_font
        cat_cell.fill = category_fill
        cat_cell.alignment = left_align
        cat_cell.border = thin_border
        for c in range(2, 7):
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 25
        row += 1

    # データ行
    values = [label, r1, r2, o1, o2, analysis]
    ws.row_dimensions[row].height = 60
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = cell_font
        cell.alignment = left_align if col in [1, 6] else center_align
        cell.border = thin_border

        # 当社列のハイライト
        if col in [4, 5]:
            if advantage == "good":
                cell.fill = highlight_fill
            elif advantage == "bad":
                cell.fill = weak_fill

    # 分析列の色
    analysis_cell = ws.cell(row=row, column=6)
    if advantage == "good":
        analysis_cell.fill = highlight_fill
    elif advantage == "bad":
        analysis_cell.fill = weak_fill

    row += 1

# === サマリーセクション ===
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
summary_title = ws.cell(row=row, column=1, value="■ 総合サマリー")
summary_title.font = Font(name="Meiryo UI", size=12, bold=True, color="FFFFFF")
summary_title.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
summary_title.alignment = center_align
for c in range(2, 7):
    ws.cell(row=row, column=c).border = thin_border
ws.row_dimensions[row].height = 30
row += 1

summaries = [
    ("当社の強み（活かすべき点）",
     "① 当日発送（12時まで） → 急ぎのギフト需要に圧倒的優位\n"
     "② アクリルスタンド付属 → 壁掛け不要の手軽さで差別化\n"
     "③ ギフト用途の提案充実 → 結婚・出産・記念日ニーズを取り込み\n"
     "④ 90日保証 → 安心感の訴求\n"
     "⑤ 専門店ブランド「名入れ屋」 → 信頼性の訴求",
     highlight_fill),
    ("当社の課題（改善すべき点）",
     "① 価格競争力 → ライバルの「1000円ポッキリ」に対し約2倍の価格差\n"
     "② サイズ展開の少なさ → ライバル8種 vs 当社3種\n"
     "③ レビュー数の不足 → ライバル合計1,200件超に対し当社は未確認\n"
     "④ セット販売未対応 → 客単価向上の機会損失\n"
     "⑤ 価格訴求コピーの不足 → キャッチーな価格表現が必要",
     weak_fill),
    ("推奨アクション",
     "① 「当日届く！スマホで簡単キャンバスプリント」等のUSP明確化\n"
     "② エントリー価格帯（1,000円台前半）の商品追加検討\n"
     "③ 2枚・3枚セット割引の導入で客単価UP\n"
     "④ レビュー獲得キャンペーン（購入後フォローメール等）\n"
     "⑤ ギフトラッピング・メッセージカード等のオプション拡充\n"
     "⑥ 正方形サイズの追加でインテリア需要に対応",
     PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")),
]

for title, content, fill in summaries:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value=title)
    title_cell.font = Font(name="Meiryo UI", size=10, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    title_cell.fill = fill
    title_cell.border = thin_border
    content_cell = ws.cell(row=row, column=2, value=content)
    content_cell.font = cell_font
    content_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    content_cell.fill = fill
    content_cell.border = thin_border
    for c in range(3, 7):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 110
    row += 1

# === 凡例 ===
row += 1
ws.cell(row=row, column=1, value="【凡例】").font = Font(name="Meiryo UI", size=9, bold=True)
row += 1
legend_good = ws.cell(row=row, column=1, value="  緑色 = 当社優位項目")
legend_good.font = Font(name="Meiryo UI", size=9)
legend_good.fill = highlight_fill
row += 1
legend_bad = ws.cell(row=row, column=1, value="  赤色 = 当社課題項目")
legend_bad.font = Font(name="Meiryo UI", size=9)
legend_bad.fill = weak_fill

# 保存
output_path = r"C:\Users\siwah\claude code\2026.3.4_キャンバスプリント不振挽回思案\ライバル商品比較表.xlsx"
wb.save(output_path)
print(f"保存完了: {output_path}")
